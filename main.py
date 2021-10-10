import openpyxl

inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file["Sheet1"]

# defining the dictionary
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# product_row.max_row is the total number of rows in that specific sheet
# print(product_list.max_row)
# we use range because to use for loop we have to have a range of numbers
# and inorder to skip the first row since it has the title we use 2 'range(2' otherwise we would just leave it out
# with range the last value is not considered so we have to add 1 '+ 1'
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    # adding the last column in the cell
    inventory_price = product_list.cell(product_row, 5)

    # calculation for number of products per supplier
    # check if supplier already in the dictionary then add 1
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        # here we are assigning a value to key supplier
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation for total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic calculation for inventory under 10
    # checks to see which values are about to get out of stock
    if inventory < 10:
        products_under_10_inv[int(product_num)] = inventory

    # logic to add total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# generating a new file
inv_file.save("inventory_with_total_value.xlsx")


# Output
# {25: 7.0, 30: 6.0, 74: 2.0}
# {'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14}
# {'AAA Company': 10969059.95, 'BBB Company': 2375499.47, 'CCC Company': 8114363.62}
